"""
parser.py

Reads a FAST-NUCES-style timetable Excel file (one sheet per day, a grid of
classroom x time-slot) and converts it into a flat list of ClassSession
objects that the rest of the app can work with.

Expected sheet layout (rows are 1-indexed as seen in Excel):
    Row 1: day name in col A (e.g. "Monday")
    Row 2: "Slots" in col A, then slot numbers (1, 2, 3, ...) in following cols
    Row 3: "Venues/time" in col A, then time ranges (e.g. "08:00-8:55")
    Row 4: "CLASSROOMS" header (no data)
    Row 5 onward: classroom name in col A, then one cell per slot containing
                  either nothing (free) or "COURSE SECTION\nTEACHER NAME"

A "LABS" header row may appear partway down (just a section label, not data)
and is skipped automatically because it has no slot data in its row.

Lab sessions are 3 hours long but the source file only marks the *starting*
slot — slots 2 and 3 of a lab show up as blank even though the room and
section are occupied. This module detects lab entries (by "lab" appearing
in the course or section text) and auto-expands them across the following
two slots.

A handful of rows (~2%) have no \\n separating the section code from the
teacher name, using a long run of spaces instead, and some of those also
have an extra label (e.g. "Net", "Lab") stuck in front of the section code
in a way that can't be reliably auto-corrected. Those rows are returned
separately as "flagged" rather than guessed at, so they can be reviewed by
hand rather than silently producing a wrong answer.
"""

from __future__ import annotations
from dataclasses import dataclass
import re
import openpyxl
from openpyxl.workbook.workbook import Workbook

from overrides import MANUAL_OVERRIDES
from course_aliases import normalize_course
from teacher_aliases import normalize_teacher


@dataclass(frozen=True)
class ClassSession:
    """One scheduled class: a specific course+section meeting at a specific
    day/slot/classroom with a specific teacher."""
    day: str
    slot: int
    time_range: str
    classroom: str
    course: str
    section: str
    teacher: str
    is_lab: bool = False

    @property
    def course_section(self) -> str:
        """Unique key for a specific section of a course, e.g. 'DAA BCS-5C'."""
        return f"{self.course} {self.section}"


@dataclass(frozen=True)
class FlaggedCell:
    """A cell that looked like it might be a class but couldn't be parsed
    with confidence — surfaced for manual review instead of guessed at."""
    day: str
    slot: int
    classroom: str
    raw_text: str
    reason: str


# Rows that are headers/labels rather than actual classroom data.
_NON_CLASSROOM_LABELS = {"slots", "venues/time", "classrooms", "labs"}

# Junk entries that look like a class but aren't (event reservations, etc.)
_JUNK_PATTERNS = [
    re.compile(r"reserved", re.IGNORECASE),
    re.compile(r"nu fest", re.IGNORECASE),
    re.compile(r"jumma|namaz|prayer break", re.IGNORECASE),
]

# Matches a section code anywhere in a string: 2-6 uppercase letters,
# optionally a dash (with optional spaces around it) or just a space,
# then a digit, optionally followed by a letter. Anchored as a search
# (not just a prefix match) because the course name itself can be
# multiple words (e.g. "B. Math 1", "Basic Eco", "DC Net") — the section
# code is whatever comes right before this pattern, however many words
# that is.
_SECTION_PATTERN = re.compile(r"\b[A-Z]{2,6}\s*-?\s*\d[A-Za-z]?\b")


def _split_course_and_section(header: str) -> tuple[str, str] | None:
    """
    Split a "COURSE SECTION" header into (course, section) by finding
    where the section-code pattern starts, rather than assuming the
    course name is always exactly one word.

    e.g. "Basic Eco BSBA-5A" -> ("Basic Eco", "BSBA-5A")
         "B. Math 1 BSFT-1A" -> ("B. Math 1", "BSFT-1A")
         "DAA BCS-5C"        -> ("DAA", "BCS-5C")

    Returns None if no section-code-shaped substring is found at all.
    """
    match = _SECTION_PATTERN.search(header)
    if match is None:
        return None
    course = header[: match.start()].strip().rstrip("-").strip()
    section = header[match.start():].strip()
    if not course or not section:
        return None
    return course, section

# How many consecutive slots a lab occupies (the source file only labels
# the first one).
_LAB_DURATION_SLOTS = 3


def _is_junk(text: str) -> bool:
    return any(pattern.search(text) for pattern in _JUNK_PATTERNS)


def _clean(text: str) -> str:
    """Collapse whitespace and strip — the source data has inconsistent
    spacing (e.g. double spaces, leading/trailing newlines)."""
    return re.sub(r"\s+", " ", text).strip()


def _is_lab(course: str, section: str) -> bool:
    return "lab" in course.lower() or "lab" in section.lower()


class _ParseResult:
    """Internal sentinel distinguishing 'nothing here' from 'something here
    but we're not confident enough to parse it automatically'."""
    EMPTY = object()
    FLAGGED = object()


def _parse_cell(raw, day: str, slot: int, classroom: str):
    """
    Parse a single timetable cell.

    Returns one of:
      - None                  -> cell was genuinely empty or junk, skip it
      - (course, section, teacher, is_lab)  -> parsed with confidence
      - FlaggedCell            -> looked like a class but couldn't be split
                                   reliably; needs manual review
    """
    if raw is None:
        return None

    text = str(raw).strip()
    if not text:
        return None
    if _is_junk(text):
        return None

    # Check the manual override list first — these are known-messy cells
    # that were reviewed and corrected by hand, regardless of which
    # splitting path they'd otherwise hit.
    if text in MANUAL_OVERRIDES:
        course, section, teacher = MANUAL_OVERRIDES[text]
        course = normalize_course(course)
        teacher = normalize_teacher(teacher)
        return course, section, teacher, _is_lab(course, section)

    # Split "COURSE SECTION" (first line) from "TEACHER" (rest)
    parts = text.split("\n", 1)
    header = _clean(parts[0])
    teacher = _clean(parts[1]) if len(parts) > 1 else ""

    # header is something like "DAA BCS-5C", "Basic Eco BSBA-5A", or
    # "NPS BCY-5A (A,B)" — the course name itself can be multiple words,
    # so we anchor on where the section-code pattern starts rather than
    # assuming the course is always exactly the first word.
    split = _split_course_and_section(header)
    if split is None:
        # No section-code-shaped substring anywhere — not a real class.
        return None

    course, section = split

    # If there was no \n at all, "teacher" will be empty and the real
    # teacher name is still stuck onto the end of "section". The header
    # split above is greedy (matches the FIRST section-code-shaped
    # substring), so "section" here may actually be "BSEE-5A    Engr.
    # Sadaf Ayesha" — try splitting on the now-known section boundary.
    if not teacher:
        # Re-find where the section code actually ends within `section`,
        # since `_split_course_and_section` only told us where it starts.
        match = _SECTION_PATTERN.search(section)
        if match and match.start() == 0:
            leftover = section[match.end():].strip()
            if leftover:
                section, teacher = match.group(0), leftover
            else:
                return FlaggedCell(
                    day, slot, classroom, text,
                    "no newline separator and no clear teacher name after section code",
                )
        else:
            return FlaggedCell(
                day, slot, classroom, text,
                "no newline separator and section code pattern not found at expected position",
            )

    course = normalize_course(course)
    teacher = normalize_teacher(teacher)
    return course, section, teacher, _is_lab(course, section)


def parse_day_sheet(
    ws, day_name: str
) -> tuple[list[ClassSession], list[FlaggedCell]]:
    """Parse a single day's worksheet into class sessions + flagged cells."""
    sessions: list[ClassSession] = []
    flagged: list[FlaggedCell] = []

    # Row 2 holds slot numbers, row 3 holds the matching time ranges.
    slot_row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    time_row = next(ws.iter_rows(min_row=3, max_row=3, values_only=True))

    slot_numbers = [int(v) for v in slot_row[1:] if v is not None]
    time_ranges = [str(v) for v in time_row[1 : 1 + len(slot_numbers)]]
    max_slot = max(slot_numbers)

    for row in ws.iter_rows(min_row=5, values_only=True):
        classroom_raw = row[0]
        if classroom_raw is None:
            continue

        classroom = _clean(str(classroom_raw))
        if classroom.lower() in _NON_CLASSROOM_LABELS:
            continue

        for col_index, slot_number in enumerate(slot_numbers, start=1):
            if col_index >= len(row):
                break

            result = _parse_cell(row[col_index], day_name, slot_number, classroom)
            if result is None:
                continue
            if isinstance(result, FlaggedCell):
                flagged.append(result)
                continue

            course, section, teacher, is_lab = result
            duration = _LAB_DURATION_SLOTS if is_lab else 1

            # Expand across `duration` consecutive slots, capped at the
            # last slot of the day.
            for offset in range(duration):
                occupied_slot = slot_number + offset
                if occupied_slot > max_slot:
                    break
                time_idx = occupied_slot - slot_numbers[0]
                sessions.append(
                    ClassSession(
                        day=day_name,
                        slot=occupied_slot,
                        time_range=time_ranges[time_idx]
                        if 0 <= time_idx < len(time_ranges)
                        else "",
                        classroom=classroom,
                        course=course,
                        section=section,
                        teacher=teacher,
                        is_lab=is_lab,
                    )
                )

    return sessions, flagged


# Maps expected sheet names (case/whitespace vary in the source file) to a
# clean canonical day name.
_DAY_SHEET_ALIASES = {
    "monday": "Monday",
    "tuesday": "Tuesday",
    "wednesday": "Wednesday",
    "thursday": "Thursday",
    "friday": "Friday",
}


def parse_timetable(
    file_path: str,
) -> tuple[list[ClassSession], list[FlaggedCell]]:
    """Parse the full timetable workbook (all weekday sheets) into a flat
    list of ClassSession objects, plus any cells that need manual review."""
    wb: Workbook = openpyxl.load_workbook(file_path, data_only=True)

    all_sessions: list[ClassSession] = []
    all_flagged: list[FlaggedCell] = []
    for sheet_name in wb.sheetnames:
        key = sheet_name.strip().lower()
        if key not in _DAY_SHEET_ALIASES:
            continue
        day_name = _DAY_SHEET_ALIASES[key]
        sessions, flagged = parse_day_sheet(wb[sheet_name], day_name)
        all_sessions.extend(sessions)
        all_flagged.extend(flagged)

    return all_sessions, all_flagged


if __name__ == "__main__":
    import sys

    path = sys.argv[1] if len(sys.argv) > 1 else "timetable.xlsx"
    sessions, flagged = parse_timetable(path)
    print(f"Parsed {len(sessions)} class sessions ({sum(s.is_lab for s in sessions)} are lab slots).")
    print(f"Flagged {len(flagged)} cells for manual review:\n")
    for f in flagged:
        print(f"  [{f.day} slot {f.slot}] {f.classroom}: {f.raw_text!r}")
        print(f"    reason: {f.reason}")